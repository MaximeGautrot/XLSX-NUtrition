import os
import openpyxl
from unidecode import unidecode

def format_text(text):
    # Remove accents
    text_without_accents = unidecode(text)

    # Capitalize the first letter of each word
    formatted_text = text_without_accents.title()

    return formatted_text

def convert_txt_to_xls(input_file, output_file, template_file):
    # Load the template file
    workbook_template = openpyxl.load_workbook(template_file, read_only=True)

    # Create a new workbook and copy the content from the template
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet_template = workbook_template.active

    # Copy the cells from the template to the new workbook
    for row in worksheet_template.iter_rows():
        for cell in row:
            worksheet.cell(row=cell.row, column=cell.column, value=cell.value)

    # Copy the styles from the template to the new workbook
    worksheet.column_dimensions = worksheet_template.column_dimensions
    worksheet.row_dimensions = worksheet_template.row_dimensions

    with open(input_file, 'r', encoding="UTF-8") as txt_file:
        lines = txt_file.readlines()

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the header row
    worksheet.cell(row=1, column=1, value="Nom")
    worksheet.cell(row=1, column=2, value="Produits")
    worksheet.cell(row=1, column=4, value="Produits")
    worksheet.cell(row=1, column=5, value="Nombres")

    # Apply header formatting
    header_font = openpyxl.styles.Font(bold=True)
    header_fill = openpyxl.styles.PatternFill(start_color='00C050', end_color='00C050', fill_type='solid')
    thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'),
                                                 right=openpyxl.styles.borders.Side(style='thin'),
                                                 top=openpyxl.styles.borders.Side(style='thin'),
                                                 bottom=openpyxl.styles.borders.Side(style='thin'))
    for cell in worksheet[1]:
        if cell.value is not None:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

    # Initialize the starting row for data and lists to store column lengths
    row_idx = 2
    col_lengths = [len("Nom"), len("Produits")]

    for line in lines:
        line = format_text(line)
        person, commands = line.strip().split(': ')
        products = commands.split(', ')

        # Write each person and their product to a new row and update column lengths
        for i, product in enumerate(products):
            worksheet.cell(row=row_idx, column=1, value=person)
            worksheet.cell(row=row_idx, column=2, value=product)
            worksheet.cell(row=row_idx, column=4, value=product)
            worksheet.cell(row=row_idx, column=5, value=i + 1)  # Assign a number to each product
            col_lengths[0] = max(col_lengths[0], len(person))
            col_lengths[1] = max(col_lengths[1], len(product))
            row_idx += 1

    # Apply data formatting
    for row in worksheet.iter_rows(min_row=2, values_only=False):
        for cell in row:
            if cell.column_letter != 'C':
                cell.border = thin_border

    # Remove formatting from column C
    no_fill = openpyxl.styles.fills.Fill()
    for cell in worksheet['C']:
        cell.fill = no_fill
        cell.border = None

    # Adjust column widths based on the longest element in each column
    for i, col_length in enumerate(col_lengths):
        column_dim = worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)]
        column_dim.width = col_length + 2  # Add some extra space

    column_dim = worksheet.column_dimensions[openpyxl.utils.get_column_letter(4)]
    column_dim.width = col_lengths[1] + 2  # Add some extra space

    # Save the workbook to the output file
    workbook.save(output_file)

if __name__ == '__main__':
    current_directory = os.getcwd()
    input_file = 'input.txt'  # Chemin du fichier texte
    output_file = 'output.xls'  # Chemin du fichier CSV de sortie

    convert_txt_to_xls(input_file, output_file)