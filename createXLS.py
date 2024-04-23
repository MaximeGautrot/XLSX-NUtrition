import openpyxl
from unidecode import unidecode
import json
import shutil

def format_text(text):
    # Remove accents
    formatted_text = unidecode(text)

    return formatted_text

def format_text_capitilize(text):
    # Remove accents
    text_without_accents = unidecode(text)
    
    # Capitalize the first letter of each word
    formatted_text = text_without_accents.title()
    
    return formatted_text

def read_data_from_column(worksheet, column_index):
    data = []
    row_idx = 2

    while True:
        cell = worksheet.cell(row=row_idx, column=column_index)
        if cell.value is None or cell.value == '':
            break

        data.append(cell.value)
        row_idx += 1

    return data

def get_information(input_file):
    workbook_input = openpyxl.load_workbook(input_file, read_only=True)
    worksheet_input = workbook_input.active

    with open('produits.json', 'r', encoding='UTF-8') as json_file:
        dict_produits = json.load(json_file)
    
    dict_info = {}
    nb_panier_legume = 0
    nb_demi_panier_legume = 0
    nb_panier_fruit = 0
    dict_fromage_pain = {'Fromage': {}, 'Pain': {}}

    data_prenom = read_data_from_column(worksheet_input, 1)
    data_nom = read_data_from_column(worksheet_input, 2)
    data_article = read_data_from_column(worksheet_input, 5)
    data_quantite = read_data_from_column(worksheet_input, 6)
    
    for i in range(len(data_prenom)):
        key = data_nom[i] + ' ' + data_prenom[i]
        
        if(not(key in dict_info.keys())):
            dict_info[key] = []
        
        dict_info[key].append((data_article[i], data_quantite[i]))
        
        if(data_article[i] == "Panier de légumes (2,5 kg)"):
            nb_panier_legume += 1
        elif(data_article[i] == "Demi panier de légumes"):
            nb_demi_panier_legume += 1
        elif(data_article[i] == "Fruits"):
            nb_panier_fruit += 1
            dict_info[key].pop()
            dict_info[key].append(("Panier de fruits", data_quantite[i]))
        elif(data_article[i] in dict_produits.keys()):
            if(data_article[i] in dict_fromage_pain[dict_produits[data_article[i]]].keys()):
                dict_fromage_pain[dict_produits[data_article[i]]][data_article[i]] += 1
            else:
                dict_fromage_pain[dict_produits[data_article[i]]][data_article[i]] = 1
    
    return dict_info, nb_panier_legume, nb_demi_panier_legume, nb_panier_fruit, dict_fromage_pain
    
    
def create_xls(input_file, output_file, template_file):
    shutil.copyfile(template_file, output_file)
    shutil.copystat(template_file, output_file)
    
    # Load the template copy file
    workbook = openpyxl.load_workbook(output_file, read_only=False)
    worksheet = workbook.active

    # Write the header row
    worksheet.cell(row=1, column=1, value="Nom")
    worksheet.cell(row=1, column=2, value="Produits")
    worksheet.cell(row=1, column=4, value="Produits")
    worksheet.cell(row=1, column=5, value="Nombres")

    # Apply header formatting
    header_font = openpyxl.styles.Font(bold=True)
    header_fill = openpyxl.styles.PatternFill(start_color='00C050', end_color='00C050', fill_type='solid')
    thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'),
                                                 right=openpyxl.styles.borders.Side(style='thin'),
                                                 top=openpyxl.styles.borders.Side(style='thin'),
                                                 bottom=openpyxl.styles.borders.Side(style='thin'))
    for cell in worksheet[1]:
        if cell.value is not None:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

    # Initialize the starting row for data and lists to store column lengths
    row_idx = 2
    col_lengths = [len("Nom"), len("Produits")]
    
    dict_info, nb_panier_legume, nb_demi_panier_legume, nb_panier_fruit, dict_fromage_pain = get_information(input_file)

    for key in dict_info.keys():
        for elt0, elt1 in dict_info[key]:
            worksheet.cell(row=row_idx, column=1, value=key)
            worksheet.cell(row=row_idx, column=2, value=elt0)
            worksheet.cell(row=row_idx, column=4, value=elt0)
            worksheet.cell(row=row_idx, column=5, value=elt1)  # Assign a number to each product
            col_lengths[0] = max(col_lengths[0], len(key))
            col_lengths[1] = max(col_lengths[1], len(elt0))
            row_idx += 1
    
    column_letter_basic = ['C', 'F', 'J', 'O', 'P']
    column_letter_not_touch = ['Q', 'R']

    # Apply data formatting
    for row in worksheet.iter_rows(min_row=2, values_only=False):
        for cell in row:
            if(not((cell.column_letter in column_letter_basic) or (cell.column_letter in column_letter_not_touch))):
                cell.border = thin_border

    # Remove formatting from column C
    no_fill = openpyxl.styles.fills.Fill()
    for letter in column_letter_basic:
        for cell in worksheet[letter]:
            cell.fill = no_fill
            cell.border = None

    # Adjust column widths based on the longest element in each column
    for i, col_length in enumerate(col_lengths):
        column_dim = worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)]
        column_dim.width = col_length + 2  # Add some extra space

    column_dim = worksheet.column_dimensions[openpyxl.utils.get_column_letter(4)]
    column_dim.width = col_lengths[1] + 2  # Add some extra space
    
    worksheet.cell(row=2, column=18, value=nb_panier_fruit)
    worksheet.cell(row=3, column=18, value=nb_panier_legume)
    worksheet.cell(row=4, column=18, value=nb_demi_panier_legume)
    
    len_Q = 0

    for i, key in enumerate(dict_fromage_pain['Fromage'].keys()):
        worksheet.cell(row=8+i, column=17, value=key)
        worksheet.cell(row=8+i, column=18, value=dict_fromage_pain['Fromage'][key])
        if(len(key) > len_Q):
            len_Q = len(key)

    for i, key in enumerate(dict_fromage_pain['Pain'].keys()):
        worksheet.cell(row=17+i, column=17, value=key)
        worksheet.cell(row=17+i, column=18, value=dict_fromage_pain['Pain'][key])
        if(len(key) > len_Q):
            len_Q = len(key)
    
    column_dim = worksheet.column_dimensions[openpyxl.utils.get_column_letter(17)]
    column_dim.width = len_Q  # Add some extra space

    # Save the workbook to the output file
    workbook.save(output_file)